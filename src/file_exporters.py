"""
File export utilities for SRS Document Generator
Handles CSV, Excel, and other format exports with various encodings
"""

import csv
import json
from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class FileExporter:
    """Utility class for exporting data to various file formats"""

    def __init__(self, output_dir: Path):
        """Initialize exporter with output directory"""
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_csv(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        encoding: str = "utf-8",
        include_edge_cases: bool = False,
        edge_case_probability: float = 0.1,
    ) -> Path:
        """Export data to CSV file with specified encoding"""
        if not data:
            raise ValueError("No data provided for export")

        # Add edge cases if requested
        if include_edge_cases:
            edge_cases = self._add_edge_cases(data, edge_case_probability)
            data = data + edge_cases

        file_path = self.output_dir / f"{filename}.csv"

        # Get fieldnames from first record
        fieldnames = list(data[0].keys())

        with open(file_path, "w", newline="", encoding=encoding) as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)

        return file_path

    def export_excel(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        sheet_name: str = "Data",
        include_formatting: bool = True,
    ) -> Path:
        """Export data to Excel file (.xlsx)"""
        if not data:
            raise ValueError("No data provided for export")

        file_path = self.output_dir / f"{filename}.xlsx"

        # Create DataFrame
        df = pd.DataFrame(data)

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        if include_formatting:
            self._format_excel_worksheet(ws, df)

        wb.save(file_path)
        return file_path

    def export_excel_legacy(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        sheet_name: str = "Data",
    ) -> Path:
        """Export data to legacy Excel file (.xls)"""
        if not data:
            raise ValueError("No data provided for export")

        file_path = self.output_dir / f"{filename}.xls"

        # Create DataFrame
        df = pd.DataFrame(data)

        # Export to .xls format
        df.to_excel(
            file_path, sheet_name=sheet_name, index=False, engine="xlwt"
        )

        return file_path

    def export_json(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        encoding: str = "utf-8",
        indent: int = 2,
    ) -> Path:
        """Export data to JSON file"""
        if not data:
            raise ValueError("No data provided for export")

        file_path = self.output_dir / f"{filename}.json"

        with open(file_path, "w", encoding=encoding) as jsonfile:
            json.dump(data, jsonfile, indent=indent, ensure_ascii=False)

        return file_path

    def export_data_dictionary(
        self, fields_dict: Dict[str, str], filename: str = "data_dictionary"
    ) -> Path:
        """Export data dictionary to CSV file"""
        file_path = self.output_dir / f"{filename}.csv"

        # Convert to list of dictionaries
        data = [
            {"field_name": field, "description": description}
            for field, description in fields_dict.items()
        ]

        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            fieldnames = ["field_name", "description"]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)

        return file_path

    def _format_excel_worksheet(self, ws, df):
        """Apply formatting to Excel worksheet"""
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Apply header formatting
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

    def _add_edge_cases(
        self, data: List[Dict[str, Any]], probability: float
    ) -> List[Dict[str, Any]]:
        """Add edge cases to data for testing validation"""
        import random

        edge_cases = []

        for item in data:
            if random.random() < probability:
                # Create a copy and modify it
                edge_item = item.copy()

                # Randomly introduce edge cases
                edge_type = random.choice(
                    [
                        "empty_string",
                        "special_chars",
                        "large_number",
                        "negative_number",
                        "zero_value",
                        "very_long_string",
                        "unicode_chars",
                    ]
                )

                if edge_type == "empty_string":
                    # Make some fields empty
                    fields_to_empty = random.sample(
                        list(edge_item.keys()), random.randint(1, 3)
                    )
                    for field in fields_to_empty:
                        if isinstance(edge_item[field], str):
                            edge_item[field] = ""

                elif edge_type == "special_chars":
                    # Add special characters
                    special_chars = "!@#$%^&*()_+-=[]{}|;':\",./<>?"
                    for key, value in edge_item.items():
                        if isinstance(value, str) and random.random() < 0.3:
                            edge_item[key] = (
                                f"{value}{random.choice(special_chars)}"
                            )

                elif edge_type == "large_number":
                    # Make numeric fields very large
                    for key, value in edge_item.items():
                        if (
                            isinstance(value, (int, float))
                            and random.random() < 0.3
                        ):
                            edge_item[key] = value * random.randint(
                                1000, 10000
                            )

                elif edge_type == "negative_number":
                    # Make numeric fields negative
                    for key, value in edge_item.items():
                        if (
                            isinstance(value, (int, float))
                            and random.random() < 0.3
                        ):
                            edge_item[key] = -abs(value)

                elif edge_type == "zero_value":
                    # Make numeric fields zero
                    for key, value in edge_item.items():
                        if (
                            isinstance(value, (int, float))
                            and random.random() < 0.3
                        ):
                            edge_item[key] = 0

                elif edge_type == "very_long_string":
                    # Make string fields very long
                    for key, value in edge_item.items():
                        if isinstance(value, str) and random.random() < 0.3:
                            edge_item[key] = value * random.randint(10, 50)

                elif edge_type == "unicode_chars":
                    # Add unicode characters
                    unicode_chars = "αβγδεζηθικλμνξοπρστυφχψω"
                    for key, value in edge_item.items():
                        if isinstance(value, str) and random.random() < 0.3:
                            edge_item[key] = (
                                f"{value}{random.choice(unicode_chars)}"
                            )

                edge_cases.append(edge_item)

        return edge_cases

    def create_error_files(
        self, data: List[Dict[str, Any]], filename_prefix: str
    ) -> List[Path]:
        """Create files with intentional errors for validation testing"""
        error_files = []

        # File with missing required fields
        error_data = []
        for item in data[:10]:  # Take first 10 items
            error_item = item.copy()
            # Remove some required fields
            fields_to_remove = list(error_item.keys())[:3]
            for field in fields_to_remove:
                del error_item[field]
            error_data.append(error_item)

        error_file = self.output_dir / f"{filename_prefix}_missing_fields.csv"
        with open(error_file, "w", newline="", encoding="utf-8") as csvfile:
            if error_data:
                fieldnames = list(error_data[0].keys())
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(error_data)
        error_files.append(error_file)

        # File with invalid data types
        error_data = []
        for item in data[:10]:
            error_item = item.copy()
            # Convert numeric fields to strings with invalid values
            for key, value in error_item.items():
                if isinstance(value, (int, float)):
                    error_item[key] = "invalid_number"
            error_data.append(error_item)

        error_file = self.output_dir / f"{filename_prefix}_invalid_types.csv"
        with open(error_file, "w", newline="", encoding="utf-8") as csvfile:
            if error_data:
                fieldnames = list(error_data[0].keys())
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(error_data)
        error_files.append(error_file)

        # File with malformed CSV (extra commas, quotes)
        error_file = self.output_dir / f"{filename_prefix}_malformed.csv"
        with open(error_file, "w", newline="", encoding="utf-8") as csvfile:
            csvfile.write("field1,field2,field3\n")
            csvfile.write('"value1","value2","value3"\n')
            csvfile.write('"value1,with,commas","value2","value3"\n')
            csvfile.write('"value1","value2"with"quotes","value3"\n')
            csvfile.write('"value1","value2","value3"\n')
        error_files.append(error_file)

        return error_files
